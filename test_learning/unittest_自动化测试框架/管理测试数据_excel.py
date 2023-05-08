'''
openpyxl    --仅支持.xlsx读写操作

excel的三个对象
    1、工作簿（Workbook）
    2、表单（Sheet）
    3、单元格（Cell）

1、打开测试数据文件，生成Workbook对象                load_workbook模块
2、从工作簿中获取表单对象                           --wb['表单名']
3、从表单中获取单元格对象                           --sh(row,column)      ##openpyxl中下标以1开始

对工作簿的操作
    1、  .save(路径)          默认保存，若填写路径则为另存为效果             ##保存时注意文件不可被占用
    2、

对表单的操作
    1、  .max_row            总行数
    2、  .max_column         总列数
    3、  .rows               按行获取所有cell(生成器)                    ##需使用list获取内部信息
                                                                     ##这里计数从0开始

对单元格的操作
    1、  .value              取值
    2、  .value = 新数据      修改
    3、
'''

from openpyxl import load_workbook

# 1、打开测试数据文件，生成Workbook对象                load_workbook模块
wb = load_workbook('login_cases_data.xlsx')

# 2、从工作簿中获取表单对象                           --wb['表单名']
sheet = wb['Sheet1']

# 3、从表单中获取单元格对象                           --sell(row,colum)      ##openpyxl中下标以1开始
cell_2_2 = sheet.cell(2, 2)                                             ###row等函数从1开始计数

print(sheet.max_row)
print(sheet.max_column)
value_2_2 = cell_2_2.value
print(value_2_2)

'''
sheet.rows      按行获取到整个表单内cell对象
'''

title = []
for items in list(sheet.rows)[0]:
    title.append(items.value)

print(title)

datas = []

for items in list(sheet.rows)[1:]:
    dic = {}
    for index, cell in enumerate(items):
        dic[title[index]] = cell.value
    datas.append(dic)

print(datas)

'''
引入内置zip()函数         可以直接将两个长度相同的内容打包成可迭代对象

下面进行代码优化
'''

datas1 = []

title1 = []
for items in list(sheet.rows)[0]:
    title1.append(items.value)

for items in list(sheet.rows)[1:]:
    content = []
    for cell in items:
        content.append(cell.value)
    datas1.append(dict(zip(title1, content)))

print(datas == datas1)

'''
excel读取格式问题         只用数字、字符串

引入eval()              可以将字符串当作python语句执行
'''

datas2 = []

for items in list(sheet.rows)[1:]:
    content = []
    for index, cell in enumerate(items):
        if index == 2:
            content.append(eval(cell.value))
        else:
            content.append(cell.value)
    datas2.append(dict(zip(title1, content)))

print(datas2)
