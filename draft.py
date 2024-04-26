# from 自动化.接口自动化.test_learning.post_auto_test_了解接口.layered_design.Common.handle_excel import Handle_Of_Excel
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

root = tk.Tk()
root.withdraw()  # 隐藏主窗口

file_path = filedialog.askopenfilename()

print(file_path)

excel = load_workbook(file_path)
st = excel['1']

for i in range(st.max_row):
    cell = st._get_cell(i + 1, 1)
    content = cell.value
    if content == '':
        st._get_cell(i + 1, 1).value = '一体化部件'
    else:
        st._get_cell(i + 1, 1).value = content + '、一体化部件'

excel.save(file_path)
